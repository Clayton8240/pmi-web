"""Serviço de importação de planilhas Excel para o banco PostgreSQL."""

from datetime import datetime
from sqlalchemy.orm import Session
from app.models import CD, ItemCD, Material, Campanha, VolumePorCaixa


def importar_cds_excel(excel_path: str, db: Session, nome_campanha: str | None = None) -> dict:
    """
    Importa a planilha PMI Matrix (aba 'CDs PMI') para as tabelas cds e itens_cd.

    Formato esperado (colunas):
      A=Controle Mentor, B=Cidade, C=UF, D=Regional, E=Filial,
      F=Zona Venda, G=Desc Pacote, H=Part Number, I=Marca, J=Material, K=Total

    O nome da campanha é lido automaticamente da coluna G (Descrição Pacote) da
    primeira linha válida, ou pode ser sobrescrito via `nome_campanha`.
    Cada importação cria uma nova campanha e encerra a(s) anterior(es) ativa(s).

    Retorna dict com estatísticas.
    """
    try:
        import openpyxl
    except ImportError:
        return {"erro": "openpyxl não instalado"}

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    # Detecta aba com CDs
    sheet = None
    for name in wb.sheetnames:
        if "cd" in name.lower():
            sheet = wb[name]
            break
    if sheet is None:
        sheet = wb.active

    # Detecta linha de cabeçalho
    header_row = 1
    for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        vals = [str(v).lower().strip() if v else "" for v in row]
        if any("controle" in v or "mentor" in v for v in vals):
            header_row = i
            break

    # Extrai nome da campanha da coluna G (Descrição Pacote) da 1ª linha válida
    nome_detectado: str | None = None
    for row in sheet.iter_rows(min_row=header_row + 1, max_row=header_row + 30, values_only=True):
        if row[0] is not None:
            try:
                int(row[0])
                val = str(row[6] or "").strip() if len(row) > 6 else ""
                if val:
                    nome_detectado = val
                break
            except (ValueError, TypeError):
                continue

    nome_final = (nome_campanha or nome_detectado or
                  f"Importação {datetime.now().strftime('%d/%m/%Y %H:%M')}")

    # Cria nova campanha e encerra as anteriores ativas
    db.query(Campanha).filter(Campanha.status == "ativa").update({"status": "encerrada"})
    campanha = Campanha(nome=nome_final, status="ativa")
    db.add(campanha)
    db.flush()
    campanha_id = campanha.id

    cds_inseridos = 0
    cds_atualizados = 0
    itens_inseridos = 0
    erros = []

    current_cd_id = None
    current_cd_data: dict = {}

    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        controle = row[0]
        controle_str = "" if controle is None else str(controle).strip()
        if controle_str:
            try:
                current_cd_id = int(float(controle_str))
            except (ValueError, TypeError):
                # Linha com texto na coluna A (ex.: cabeçalho repetido) — ignora
                continue

            current_cd_data = {
                "id": current_cd_id,
                "cidade": str(row[1] or "").strip(),
                "uf": str(row[2] or "").strip(),
                "regional": str(row[3] or "").strip(),
                "filial": str(row[4] or "").strip(),
                "zona_venda": str(row[5] or "").strip(),
                "descricao_pacote": str(row[6] or "").strip(),
                "cnpj": str(current_cd_id),
            }

            existing = db.get(CD, current_cd_id)
            if existing:
                for k, v in current_cd_data.items():
                    if k != "id":
                        setattr(existing, k, v)
                cds_atualizados += 1
            else:
                db.add(CD(**current_cd_data))
                cds_inseridos += 1
            db.flush()

        if current_cd_id is None:
            continue

        # Item
        part_number = str(row[7] or "").strip() if len(row) > 7 else ""
        if not part_number or part_number.lower() in ("", "none", "part number"):
            continue

        marca = str(row[8] or "").strip() if len(row) > 8 else ""
        descricao = str(row[9] or "").strip() if len(row) > 9 else ""
        try:
            qtde = int(row[10]) if len(row) > 10 and row[10] is not None else 0
        except (ValueError, TypeError):
            qtde = 0

        # Vincula ao catálogo; cria o material automaticamente se não existir
        material = (
            db.query(Material)
            .filter(Material.part_number.ilike(part_number))
            .first()
        )
        if not material:
            material = Material(
                part_number=part_number.strip().upper(),
                descricao=descricao or part_number,
                marca=marca,
                unidade="UN",
            )
            db.add(material)
            db.flush()
        else:
            # Reativa se estava desativado (usuário pode ter excluído antes
            # e a planilha trouxe ele de volta — precisa aparecer no catálogo)
            if not material.ativo:
                material.ativo = True
            # Mantém descrição/marca do catálogo atualizadas quando vierem preenchidas
            if descricao and not material.descricao:
                material.descricao = descricao
            if marca and not material.marca:
                material.marca = marca

        db.add(ItemCD(
            cd_id=current_cd_id,
            campanha_id=campanha_id,
            material_id=material.id,
            part_number=part_number,
            marca=marca,
            descricao=descricao,
            qtde=qtde,
        ))
        itens_inseridos += 1

    db.commit()

    # Desativa capacidades anteriores dos materiais desta campanha para
    # forçar o usuário a informar as quantidades por pacote a cada nova
    # importação (cada lote pode ter quantidades diferentes).
    mat_ids = [
        r[0] for r in db.query(ItemCD.material_id)
        .filter(ItemCD.campanha_id == campanha_id)
        .filter(ItemCD.material_id.isnot(None))
        .distinct()
        .all()
    ]
    if mat_ids:
        db.query(VolumePorCaixa).filter(
            VolumePorCaixa.material_id.in_(mat_ids),
            VolumePorCaixa.ativo == True,  # noqa: E712
        ).update({"ativo": False}, synchronize_session=False)
        db.commit()

    return {
        "cds_inseridos": cds_inseridos,
        "cds_atualizados": cds_atualizados,
        "itens_inseridos": itens_inseridos,
        "erros": erros,
        "campanha_id": campanha_id,
        "campanha_nome": nome_final,
    }
